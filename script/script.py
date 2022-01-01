from openpyxl import load_workbook
from datetime import date

input_file = '../excel-input-data/Labor Report 2022.xlsx'
input_text_file = '../edit-this-text-file.txt'

#create the dictionary container for the values in edited input text file
d = {}
lis = []
with open(input_text_file) as f:
    for line in f: 
        if line.lstrip().startswith('#'):
            break
        else:
            obj = line.replace('\n', '')
            (key, val) = obj.split(':')
            d[key] = val

    

#set dictionary values as variables for workbook later. 
Name = d.get('Full Name')
a_number = d.get('A Number')
pos_number = d.get('Posistion Number')
title = d.get('Title')
department = d.get('Department')
org_number = d.get('Organization Number')
signature = d.get('Signature')

#get the workbook
wb = load_workbook(input_file)

#set the worksheet we are interested in
print('Here are the active workbook sheets in {}:\n\n'.format(input_file), wb.sheetnames)
sheet = input(str('Which sheet would you like to input data using {} as the template? \n\n Enter the yyyy-dd format:'.format(input_text_file)))

ws = wb[sheet]
#set variables for the particular worksheet 
ws['E6'] = Name
ws['J6'] = a_number
ws['O6'] = pos_number
ws['R6'] = title
ws['E8'] = department
ws['C15'] = org_number
ws['D30'] = signature
ws['R30'] = date.today().strftime('%d/%m/%y')
        
#save wb to outputfile location
output_file = '../excel-input-data/Labor Report 2022.xlsx'

wb.save(output_file)